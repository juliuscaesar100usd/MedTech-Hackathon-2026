"""Loader for the real organizer service catalog.

Source: ``data/catalog/real_catalog.xlsx``, sheet ``"Справочник услуг"`` with
columns ``ID`` (specialty id), ``Специальность`` (specialty name), ``Code``
(service code), ``Name_ru`` (service name), ``TarificatrCode`` (ICD-like code).

The sheet holds 1281 service×specialty rows. A single service (by ``Code``) is
offered by several specialties, so the rows dedup to far fewer ``Service`` rows
plus one ``ServiceSpecialty`` row per source row.

Data-quality note (real file): the ``Code`` column is a worksheet formula
``=VLOOKUP(Name_ru, <table>, 2, FALSE)`` whose lookup table was deleted, so for
the lab/diagnostic block (~720 rows) ``Code`` evaluates to ``#REF!`` and the
real code is unrecoverable. Those services are still loaded in full, with
``code = NULL`` (deduped by name); the matcher can reach them by name/synonym.
Coded rows keep their integer ``Code`` as the unique key.

The loader is idempotent: re-running upserts ``Service`` rows (by ``code`` for
coded services, by name among code-less services for the rest) and inserts each
``(service, specialty)`` pair at most once.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from .config import BACKEND_ROOT
from .models import Service, ServiceSpecialty

# Sheet + canonical headers of the real catalog.
REAL_CATALOG_SHEET = "Справочник услуг"
_HEADERS = ("ID", "Специальность", "Code", "Name_ru", "TarificatrCode")

# Default on-disk location (repo_root/data/catalog/real_catalog.xlsx).
REAL_CATALOG_DEFAULT_PATH = BACKEND_ROOT.parent / "data" / "catalog" / "real_catalog.xlsx"


# --------------------------------------------------------------------------- #
# Cell coercion helpers
# --------------------------------------------------------------------------- #
def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _clean_str(value: Any) -> str | None:
    if _blank(value):
        return None
    return str(value).strip()


def _to_int(value: Any) -> int | None:
    """Coerce a cell to int, or None for blanks / non-numeric (e.g. ``#REF!``)."""
    if _blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return None


# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #
def _header_index(path: Path) -> dict[str, int] | None:
    """Return {canonical_header: column_index} if this looks like the real
    catalog (sheet present, Code + Name_ru + Специальность columns), else None.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if REAL_CATALOG_SHEET not in wb.sheetnames:
            return None
        ws = wb[REAL_CATALOG_SHEET]
        try:
            header = next(ws.iter_rows(values_only=True))
        except StopIteration:
            return None
        idx: dict[str, int] = {}
        for i, cell in enumerate(header):
            name = _clean_str(cell)
            if name in _HEADERS:
                idx[name] = i
        required = {"Специальность", "Code", "Name_ru"}
        return idx if required.issubset(idx) else None
    finally:
        wb.close()


def is_real_catalog(path: str | Path) -> bool:
    """True if ``path`` is the real organizer catalog (.xlsx with the expected
    sheet/columns). Safe to call on any upload — returns False on non-xlsx or
    unreadable files rather than raising.
    """
    p = Path(path)
    if p.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        return _header_index(p) is not None
    except Exception:
        return False


def read_real_catalog(path: str | Path) -> list[dict]:
    """Parse the real catalog into one dict per *valid* source row.

    Keys: ``code`` (int|None), ``specialty`` (str), ``specialty_id`` (int|None),
    ``service_name`` (str), ``icd_code`` (str|None). Rows missing a service name
    or specialty are skipped (the trailing blank rows in the source file).
    """
    from openpyxl import load_workbook

    p = Path(path)
    idx = _header_index(p)
    if idx is None:
        raise ValueError(
            f"{p.name!r} is not the real catalog "
            f'(expected sheet "{REAL_CATALOG_SHEET}" with Code/Name_ru/Специальность).'
        )

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb[REAL_CATALOG_SHEET]
        rows = ws.iter_rows(values_only=True)
        next(rows)  # skip header
        out: list[dict] = []
        for raw in rows:
            def col(name: str) -> Any:
                i = idx[name]
                return raw[i] if i < len(raw) else None

            service_name = _clean_str(col("Name_ru"))
            specialty = _clean_str(col("Специальность"))
            if not service_name or not specialty:
                continue  # blank/sentinel row
            out.append(
                {
                    "code": _to_int(col("Code")),
                    "specialty": specialty,
                    "specialty_id": _to_int(col("ID")),
                    "service_name": service_name,
                    "icd_code": _clean_str(col("TarificatrCode")),
                }
            )
        return out
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Loading (idempotent upsert)
# --------------------------------------------------------------------------- #
def load_real_catalog(db: Session, path: str | Path = REAL_CATALOG_DEFAULT_PATH) -> dict:
    """Idempotently load the real catalog into Service + ServiceSpecialty.

    Service identity:
      * coded rows  -> keyed by integer ``code`` (unique),
      * code-less rows -> keyed by ``service_name`` among code-less services.

    Returns a counts dict (see keys below). Safe to re-run: existing services
    are refreshed and each (service, specialty) pair is inserted at most once.
    """
    rows = read_real_catalog(path)

    # --- existing-state indexes (so re-runs upsert instead of duplicating) ---
    by_code: dict[int, Service] = {
        s.code: s
        for s in db.query(Service).filter(Service.code.isnot(None)).all()
    }
    by_name_uncoded: dict[str, Service] = {
        (s.service_name or "").strip().lower(): s
        for s in db.query(Service).filter(Service.code.is_(None)).all()
    }

    created = 0

    def resolve_service(row: dict) -> Service:
        nonlocal created
        code = row["code"]
        name = row["service_name"]
        icd = row["icd_code"]
        if code is not None:
            svc = by_code.get(code)
            if svc is None:
                svc = Service(code=code)
                db.add(svc)
                by_code[code] = svc
                created += 1
        else:
            key = name.strip().lower()
            svc = by_name_uncoded.get(key)
            if svc is None:
                svc = Service(code=None)
                db.add(svc)
                by_name_uncoded[key] = svc
                created += 1
        # Refresh descriptive fields (idempotent).
        svc.service_name = name
        svc.synonyms = [name]            # seed synonyms with Name_ru (per contract)
        if icd:
            svc.icd_code = icd
        svc.is_active = True
        return svc

    # Pass 1: ensure every Service row exists, then flush to get service_ids.
    for row in rows:
        resolve_service(row)
    db.flush()

    # Pass 2: link each (service, specialty) pair, deduped.
    existing_pairs: set[tuple[str, str]] = {
        (ss.service_id, ss.specialty) for ss in db.query(ServiceSpecialty).all()
    }
    links_created = 0
    for row in rows:
        svc = resolve_service(row)  # already cached; no new inserts here
        pair = (svc.service_id, row["specialty"])
        if pair in existing_pairs:
            continue
        existing_pairs.add(pair)
        db.add(
            ServiceSpecialty(
                service_id=svc.service_id,
                specialty=row["specialty"],
                specialty_id=row["specialty_id"],
            )
        )
        links_created += 1

    db.commit()

    # --- final counts (authoritative, read back from the DB) ---
    services_total = db.query(Service).count()
    services_coded = db.query(Service).filter(Service.code.isnot(None)).count()
    specialties_distinct = db.query(ServiceSpecialty.specialty).distinct().count()
    service_specialties = db.query(ServiceSpecialty).count()

    return {
        "rows_read": len(rows),
        "services_created": created,
        "links_created": links_created,
        "services_total": services_total,
        "services_coded": services_coded,
        "services_uncoded": services_total - services_coded,
        "service_specialties": service_specialties,
        "specialties": specialties_distinct,
    }
