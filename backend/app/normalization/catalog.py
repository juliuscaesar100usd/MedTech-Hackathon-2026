"""Loading / seeding of the target service catalog.

Supports two source formats by extension:
  * ``.xlsx`` — read with openpyxl (first sheet, first row = header),
  * ``.json`` — a list of objects (or ``{"services": [...]}``).

Column names are matched fuzzily to the canonical keys so the same loader
handles Russian or English headers (e.g. "Наименование услуги" / "name").
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from ..models import Service

# --------------------------------------------------------------------------- #
# Header aliases -> canonical key.                                             #
# --------------------------------------------------------------------------- #
_COLUMN_ALIASES: dict[str, set[str]] = {
    "service_name": {
        "service_name", "name", "service", "title",
        "наименование", "наименование услуги", "услуга", "название",
        "атауы", "қызмет",
    },
    "synonyms": {
        "synonyms", "synonym", "aliases", "alias",
        "синонимы", "синоним", "псевдонимы",
    },
    "category": {
        "category", "group", "section",
        "категория", "группа", "раздел", "санат",
    },
    # Optional deeper levels for an N-level hierarchy (schema contract §2).
    "subcategory": {
        "subcategory", "sub_category", "subgroup", "sub-group",
        "подкатегория", "подраздел", "категория2", "подгруппа",
    },
    "subsubcategory": {
        "subsubcategory", "sub_subcategory", "subsubgroup",
        "подподкатегория", "категория3",
    },
    # An explicit full path, either a list/JSON column or a delimited string.
    "category_path": {
        "category_path", "categorypath", "path", "путь", "иерархия",
    },
    "icd_code": {
        "icd_code", "icd", "icd10", "icd-10", "code", "mkb",
        "мкб", "мкб-10", "код", "код мкб",
    },
}

# Path delimiters for a path-delimited ``category``/``category_path`` string
# ('>' / '/' / '»' per the contract, plus the single-angle '›').
_PATH_SPLIT_RE = re.compile(r"\s*(?:>|/|»|›)\s*")


def _parse_path(value: Any) -> list[str]:
    """Coerce a path cell (list, or a '>'/'/'/'»'-delimited string) to list[str]."""
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if not text:
        return []
    return [p.strip() for p in _PATH_SPLIT_RE.split(text) if p.strip()]


def _canon_header(raw: str) -> str | None:
    """Map an arbitrary header cell to a canonical key, or None if unknown."""
    key = str(raw or "").strip().lower()
    if not key:
        return None
    for canon, aliases in _COLUMN_ALIASES.items():
        if key in aliases:
            return canon
    return None


def _split_synonyms(value: Any) -> list[str]:
    """Coerce a synonyms cell into a clean ``list[str]``."""
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        items = [str(v) for v in value]
    else:
        text = str(value)
        # Split on ';' or ',' (the two conventions seen in real price lists).
        parts: list[str] = []
        for chunk in text.replace(";", ",").split(","):
            parts.append(chunk)
        items = parts
    out: list[str] = []
    seen: set[str] = set()
    for it in items:
        s = it.strip()
        if s and s.lower() not in seen:
            seen.add(s.lower())
            out.append(s)
    return out


def _category_path(row: dict[str, Any]) -> list[str]:
    """Resolve the OUTER→INNER category path from any supported input form:

    (a) a path-delimited ``category`` string ("Лаборатория > Анализ крови > …"),
    (b) explicit ``category`` + ``subcategory`` (+ ``subsubcategory``) columns,
    (c) a ``category_path`` list/JSON column or delimited string.

    Form (c) wins when present; otherwise ``category`` is split on path
    delimiters and any explicit sub-level columns are appended.
    """
    # (c) explicit category_path column (list or delimited string).
    path = _parse_path(row.get("category_path"))
    if path:
        return path
    # (a) category may itself be a delimited path.
    path = _parse_path(row.get("category"))
    # (b) append explicit deeper-level columns (each may also be delimited).
    for key in ("subcategory", "subsubcategory"):
        path.extend(_parse_path(row.get(key)))
    return path


def _row_to_item(row: dict[str, Any]) -> dict | None:
    """Normalize a raw dict (already keyed by canonical names) to an item."""
    name = str(row.get("service_name") or "").strip()
    if not name:
        return None
    path = _category_path(row)
    # ``category`` stays the top level for back-compat + the dashboard rollup.
    category = path[0] if path else None
    return {
        "service_name": name,
        "synonyms": _split_synonyms(row.get("synonyms")),
        "category": category,
        "category_path": path or None,
        "icd_code": (str(row["icd_code"]).strip() or None)
        if row.get("icd_code") not in (None, "")
        else None,
    }


def _load_json(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("services") or data.get("items") or []
    items: list[dict] = []
    for obj in data:
        if not isinstance(obj, dict):
            continue
        # Re-key any aliased columns to canonical keys.
        canon: dict[str, Any] = {}
        for k, v in obj.items():
            ck = _canon_header(k) or k
            canon[ck] = v
        it = _row_to_item(canon)
        if it:
            items.append(it)
    return items


def _load_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return []
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header):
        canon = _canon_header(cell) if cell is not None else None
        if canon:
            col_map[idx] = canon
    items: list[dict] = []
    for raw in rows:
        record: dict[str, Any] = {}
        for idx, canon in col_map.items():
            if idx < len(raw):
                record[canon] = raw[idx]
        it = _row_to_item(record)
        if it:
            items.append(it)
    wb.close()
    return items


def load_catalog_from_file(path: str | Path) -> list[dict]:
    """Read a service catalog from an .xlsx or .json file.

    Returns a list of dicts with keys: service_name, synonyms(list[str]),
    category(str|None), icd_code(str|None).
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".json":
        return _load_json(p)
    if ext in {".xlsx", ".xlsm"}:
        return _load_xlsx(p)
    raise ValueError(f"Unsupported catalog file type: {ext!r} ({p.name})")


def seed_services(db: Session, items: list[dict]) -> int:
    """Upsert Service rows, matching by ``service_name``.

    For existing rows, synonyms/category/icd_code are refreshed and the row is
    re-activated. Returns the number of rows inserted or updated.
    """
    if not items:
        return 0
    existing = {s.service_name: s for s in db.query(Service).all()}
    n = 0
    for it in items:
        name = it["service_name"]
        svc = existing.get(name)
        path = it.get("category_path")
        path = list(path) if path else None
        if svc is None:
            svc = Service(
                service_name=name,
                synonyms=list(it.get("synonyms") or []),
                category=it.get("category"),
                category_path=path,
                icd_code=it.get("icd_code"),
                is_active=True,
            )
            db.add(svc)
            existing[name] = svc
        else:
            svc.synonyms = list(it.get("synonyms") or [])
            svc.category = it.get("category")
            svc.category_path = path
            svc.icd_code = it.get("icd_code")
            svc.is_active = True
        n += 1
    db.commit()
    return n


def load_services(db: Session) -> list[Service]:
    """Return all active services (the matcher's source of truth)."""
    return (
        db.query(Service)
        .filter(Service.is_active.is_(True))
        .order_by(Service.service_name)
        .all()
    )
